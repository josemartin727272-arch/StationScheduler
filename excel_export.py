"""
Export weekly schedule to a styled Excel file.
"""
import io
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Color palette
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ODD = "EBF3FB"
COLOR_ROW_EVEN = "FFFFFF"
COLOR_IL_BG = "D6E4F0"
COLOR_PE_BG = "FAD7A0"
COLOR_YELLOW = "FFD700"
COLOR_EMB_M = "A9DFBF"
COLOR_EMB_T = "F9E79F"
COLOR_VACATION = "F1948A"
COLOR_SECTION_HEADER = "2E86C1"


def export_to_excel(schedule: dict, week_start: date, lang: str = "he") -> bytes:
    """
    Build and return an Excel workbook as bytes.
    """
    from translations import t
    from schedule_logic import ROW_KEYS, EMPLOYEES_IL, EMPLOYEES_PE

    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {week_start.isoformat()}"

    day_keys = list(schedule.keys())
    num_days = len(day_keys)

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, bg=None, fg="000000", bold=False, center=False, wrap=False):
        if bg:
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        cell.font = Font(color=fg, bold=bold, size=9)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
            wrap_text=wrap,
        )
        cell.border = border

    # ── Header row: row label + 7 day columns ──────────────────────────────
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, num_days + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Row 1: dates
    ws.row_dimensions[1].height = 20
    label_cell = ws.cell(row=1, column=1, value=t("row_dates", lang))
    style_cell(label_cell, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, bold=True, center=True)
    for col_idx, dk in enumerate(day_keys, start=2):
        day = schedule[dk]
        d = day.get("date", "")
        val = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)
        c = ws.cell(row=1, column=col_idx, value=val)
        style_cell(c, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, bold=True, center=True)

    # Row 2: day names
    ws.row_dimensions[2].height = 18
    label_cell = ws.cell(row=2, column=1, value=t("row_days", lang))
    style_cell(label_cell, bg=COLOR_SECTION_HEADER, fg=COLOR_HEADER_FG, bold=True, center=True)
    day_name_keys = ["sunday", "monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]
    for col_idx, dk in enumerate(day_keys, start=2):
        d = schedule[dk].get("date")
        if hasattr(d, "weekday"):
            py_to_key = ["monday","tuesday","wednesday","thursday","friday","saturday","sunday"]
            day_name = t(py_to_key[d.weekday()], lang)
            if d.weekday() == 4:
                day_name += " (6h)"
        else:
            day_name = ""
        c = ws.cell(row=2, column=col_idx, value=day_name)
        style_cell(c, bg=COLOR_SECTION_HEADER, fg=COLOR_HEADER_FG, bold=True, center=True)

    # Rows 3-26: schedule rows
    row_display_keys = [k for k in ROW_KEYS if k not in ("dates", "days")]

    # Special formatting by row key
    row_bg_map = {
        "vacation": COLOR_VACATION,
        "vehicle_morning": None,  # handled by value
        "vehicle_noon": None,
        "udex": None,
    }

    for excel_row, rk in enumerate(row_display_keys, start=3):
        ws.row_dimensions[excel_row].height = 16
        label = t(f"row_{rk}", lang)
        label_cell = ws.cell(row=excel_row, column=1, value=label)
        is_odd = (excel_row % 2 == 1)
        row_bg = COLOR_ROW_ODD if is_odd else COLOR_ROW_EVEN
        style_cell(label_cell, bg=row_bg, bold=True)

        for col_idx, dk in enumerate(day_keys, start=2):
            day = schedule[dk]
            # vacation combines vacation_il + vacation_pe
            if rk == "vacation":
                val_il = day.get("vacation_il", "")
                val_pe = day.get("vacation_pe", "")
                parts = [p for p in [val_il, val_pe] if p]
                val = " / ".join(parts)
            else:
                val = day.get(rk, "")
            c = ws.cell(row=excel_row, column=col_idx, value=val)

            # Special coloring
            bg = row_bg
            if rk == "vacation" and val:
                bg = COLOR_VACATION
            elif rk in ("vehicle_morning", "vehicle_noon") and val == "YELLOW":
                bg = COLOR_YELLOW
            elif rk == "udex" and val == "EMB-M":
                bg = COLOR_EMB_M
            elif rk == "udex" and val == "EMB-T":
                bg = COLOR_EMB_T
            elif rk in ("emb_il", "apt_il"):
                bg = COLOR_IL_BG
            elif rk in ("emb_pe", "apt_pe"):
                bg = COLOR_PE_BG

            style_cell(c, bg=bg, center=True)

    # ── Stats sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title=t("equality_stats", lang))
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10

    # Count YELLOW
    yellow_total = sum(
        1 for dk in day_keys
        for f in ("vehicle_morning", "vehicle_noon")
        if schedule[dk].get(f) == "YELLOW"
    )
    udex_m = sum(1 for dk in day_keys if schedule[dk].get("udex") == "EMB-M")
    udex_t = sum(1 for dk in day_keys if schedule[dk].get("udex") == "EMB-T")

    ws2.append([f"YELLOW total", yellow_total, "/ 4"])
    ws2.append(["UDEX EMB-M", udex_m, "/ 3"])
    ws2.append(["UDEX EMB-T", udex_t, "/ 3"])

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
